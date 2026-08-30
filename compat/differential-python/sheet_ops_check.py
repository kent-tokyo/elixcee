#!/usr/bin/env python3
"""Differential check: elixcee's sheet management + workbook-metadata API
(P1 core 3 + remainder + P2 hidden row/col + copy_sheet + defined_names +
sheet_state + row height/column width) vs. openpyxl.

openpyxl is used PURELY as a test-only oracle here -- never a runtime dependency
of the shipped `elixcee` package (see pyproject.toml, which declares none).
Requires a `maturin develop --features python` build of elixcee and
`pip install openpyxl` in the active environment; see this directory's
README.md for the one-time setup.

Compares rename_sheet()/move_sheet()/copy_sheet()/merged_cells()/
merge_cells()/unmerge_cells()/hidden_rows()/hidden_columns()/
set_row_hidden()/set_column_hidden()/defined_names() against openpyxl's own
read of the same real fixture after a save/reload round trip. Built from
FIXTURE via elixcee.load_workbook throughout, matching the real-world usage
these tests target -- a from-scratch elixcee.Vm()'s own minimal-styles.xml
shape is covered separately by
FromScratchVmProducesAnOpenpyxlReadableStylesheet below (its bare <fill/>
rejected-on-reopen bug is fixed, see ROADMAP.md's known gaps).

sheet_state() coverage uses an openpyxl-AUTHORED fixture, not a real
Excel-authored one under compat/oracle-excel-com (none has a hidden/
veryHidden sheet), and reads only -- no elixcee save() round trip, since
elixcee has no writer support for `state="..."` yet.

row_height()/column_width() coverage also uses an openpyxl-AUTHORED fixture
(no real fixture has a genuine custom row height or column width), reads
only, same no-writer-support caveat -- a loaded row height/column width is
dropped on EVERY elixcee save today, not just sometimes (the writer
regenerates <row>/<cols> from hidden-row/column state alone).

defined_names() coverage uses FIXTURE4 (fixture1 has no <definedNames> at
all), and reads directly rather than through a save/reload round trip --
it's a passthrough-only feature this round doesn't create/modify.

Row/col insert-delete is deliberately given NO differential coverage here:
the disclosed fidelity gap (merges/hidden markers/styles/formats not shifted)
means an openpyxl comparison would correctly fail on exactly the cases worth
testing -- Rust unit + integration tests already cover the values-only
behavior that IS correct.

sort_range() also gets NO differential coverage: openpyxl has no sort
primitive of its own to compare against. Its PyO3-layer bound checks are
pinned directly (no openpyxl comparison needed) in
SortRangeAndMergeCellsRejectOversizedOrInvalidInput below.

Run standalone: `python3 compat/differential-python/sheet_ops_check.py`
"""

from __future__ import annotations

import os
import tempfile
import unittest
import warnings
import zipfile

import elixcee
import openpyxl
from openpyxl.formatting.rule import CellIsRule

FIXTURE = os.path.join(
    os.path.dirname(__file__),
    "..",
    "oracle-excel-com",
    "fixtures",
    "pristine",
    "fixture1_values_styles_merge_hidden.xlsm",
)

# The one real Excel-authored fixture with genuine <definedNames> content --
# fixture1 has none, so defined_names() coverage needs this one instead.
FIXTURE4 = os.path.join(
    os.path.dirname(__file__),
    "..",
    "oracle-excel-com",
    "fixtures",
    "pristine",
    "fixture4_hyperlink_comment_name.xlsm",
)

# The one real Excel-authored fixture with a genuine, complete <table> -- used for
# 0.16.0-A1's read-only tables() coverage.
FIXTURE3 = os.path.join(
    os.path.dirname(__file__),
    "..",
    "oracle-excel-com",
    "fixtures",
    "pristine",
    "fixture3_table_validation_conditional.xlsm",
)


class RenameSheetAgreesWithOpenpyxl(unittest.TestCase):
    def test_rename_sheet_agrees_with_openpyxl_after_a_round_trip(self):
        vm = elixcee.load_workbook(FIXTURE)
        vm.rename_sheet("Sheet1", "Renamed")

        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "out.xlsx")
            vm.save_workbook(path)

            reloaded_vm = elixcee.load_workbook(path)
            wb = openpyxl.load_workbook(path)

            self.assertEqual(wb.sheetnames, ["Renamed"])
            # elixcee's own sheet_names() convention is lowercase, unlike
            # openpyxl's display-cased sheetnames.
            self.assertEqual(reloaded_vm.sheet_names(), ["renamed"])

            ws = wb["Renamed"]
            self.assertEqual(
                [str(r) for r in ws.merged_cells.ranges],
                reloaded_vm.merged_cells(sheet="Renamed"),
            )


class CopySheetAgreesWithOpenpyxl(unittest.TestCase):
    def test_copy_sheet_agrees_with_openpyxl_after_a_round_trip(self):
        vm = elixcee.load_workbook(FIXTURE)
        vm.copy_sheet("Sheet1", "Copy")

        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "out.xlsx")
            vm.save_workbook(path)

            reloaded_vm = elixcee.load_workbook(path)
            wb = openpyxl.load_workbook(path)

            self.assertEqual(wb.sheetnames, ["Sheet1", "Copy"])
            # elixcee's own sheet_names() is alphabetically sorted, NOT
            # sheet_order/tab-position order (a pre-existing, undocumented
            # quirk unrelated to copy_sheet -- discovered while writing this
            # test). "copy" < "sheet1" alphabetically, so this is reversed
            # from wb.sheetnames' tab-position order above.
            self.assertEqual(reloaded_vm.sheet_names(), ["copy", "sheet1"])

            # The copy must carry the source's merge and hidden row/column
            # state -- both elixcee's own reload and openpyxl's independent
            # read of the same saved file must agree on it.
            ws_copy = wb["Copy"]
            self.assertEqual(
                [str(r) for r in ws_copy.merged_cells.ranges],
                reloaded_vm.merged_cells(sheet="Copy"),
            )
            self.assertIn("B1:C1", reloaded_vm.merged_cells(sheet="Copy"))
            self.assertIn(5, reloaded_vm.hidden_rows(sheet="Copy"))
            self.assertTrue(ws_copy.row_dimensions[5].hidden)


class MoveSheetAgreesWithOpenpyxl(unittest.TestCase):
    def test_move_sheet_to_the_front_matches_openpyxls_own_sheetnames(self):
        vm = elixcee.load_workbook(FIXTURE)
        vm.set_sheet("Second")  # creates a second, empty sheet
        vm.move_sheet("Second", 0)

        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "out.xlsx")
            vm.save_workbook(path)
            wb = openpyxl.load_workbook(path)
            self.assertEqual(wb.sheetnames[0], "Second")


class MergedCellsAgreesWithOpenpyxl(unittest.TestCase):
    def test_merged_cells_matches_openpyxl_on_the_real_fixture(self):
        vm = elixcee.load_workbook(FIXTURE)
        wb = openpyxl.load_workbook(FIXTURE)
        ws = wb.active
        self.assertEqual(
            vm.merged_cells(), [str(r) for r in ws.merged_cells.ranges]
        )


class MergeCellsAndUnmergeCellsAgreeWithOpenpyxl(unittest.TestCase):
    def test_a_newly_created_merge_matches_openpyxl_after_a_round_trip(self):
        vm = elixcee.load_workbook(FIXTURE)
        vm.merge_cells("D1:E1")  # non-overlapping with fixture1's pre-existing B1:C1

        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "out.xlsx")
            vm.save_workbook(path)

            reloaded_vm = elixcee.load_workbook(path)
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            self.assertEqual(
                sorted(reloaded_vm.merged_cells()),
                sorted(str(r) for r in ws.merged_cells.ranges),
            )
            self.assertIn("D1:E1", reloaded_vm.merged_cells())

    def test_removing_the_fixtures_pre_existing_merge_matches_openpyxl(self):
        vm = elixcee.load_workbook(FIXTURE)
        vm.unmerge_cells("B1:C1")

        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "out.xlsx")
            vm.save_workbook(path)

            reloaded_vm = elixcee.load_workbook(path)
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            self.assertEqual(reloaded_vm.merged_cells(), [])
            self.assertEqual(list(ws.merged_cells.ranges), [])


class HiddenRowsAndColumnsAgreeWithOpenpyxl(unittest.TestCase):
    # openpyxl's ws.row_dimensions/column_dimensions are sparse dicts,
    # populated only for rows/columns it actually parsed a <row>/<col>
    # element for -- these tests check agreement on the fixture's known
    # hidden units specifically, not an assumption that openpyxl reports
    # every hidden row/column elixcee does.
    def test_the_fixtures_pre_existing_hidden_row_and_column_match_openpyxl(self):
        vm = elixcee.load_workbook(FIXTURE)
        wb = openpyxl.load_workbook(FIXTURE)
        ws = wb.active

        self.assertIn(5, vm.hidden_rows())
        self.assertTrue(ws.row_dimensions[5].hidden)

        self.assertIn(4, vm.hidden_columns())  # column D
        self.assertTrue(ws.column_dimensions["D"].hidden)

    def test_a_newly_hidden_row_and_column_match_openpyxl_after_a_round_trip(self):
        vm = elixcee.load_workbook(FIXTURE)
        vm.set_row_hidden(20)
        vm.set_column_hidden(6)  # column F

        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "out.xlsx")
            vm.save_workbook(path)

            reloaded_vm = elixcee.load_workbook(path)
            wb = openpyxl.load_workbook(path)
            ws = wb.active

            self.assertIn(20, reloaded_vm.hidden_rows())
            self.assertTrue(ws.row_dimensions[20].hidden)

            self.assertIn(6, reloaded_vm.hidden_columns())
            self.assertTrue(ws.column_dimensions["F"].hidden)

    def test_unhiding_the_fixtures_pre_existing_hidden_row_matches_openpyxl(self):
        vm = elixcee.load_workbook(FIXTURE)
        vm.set_row_hidden(5, hidden=False)

        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "out.xlsx")
            vm.save_workbook(path)

            reloaded_vm = elixcee.load_workbook(path)
            wb = openpyxl.load_workbook(path)
            ws = wb.active

            self.assertNotIn(5, reloaded_vm.hidden_rows())
            # A <row> element with no hidden attribute at all is possible
            # (the row still exists for its cell data), so check the
            # attribute's truthiness rather than the row's mere presence.
            self.assertFalse(bool(ws.row_dimensions[5].hidden))


class DefinedNamesAgreesWithOpenpyxl(unittest.TestCase):
    def test_defined_names_matches_openpyxl_on_the_real_fixture(self):
        # FIXTURE4, not FIXTURE -- fixture1 has no <definedNames> at all.
        vm = elixcee.load_workbook(FIXTURE4)
        wb = openpyxl.load_workbook(FIXTURE4)

        elixcee_names = vm.defined_names()
        openpyxl_names = {name: dn.value for name, dn in wb.defined_names.items()}
        self.assertEqual(elixcee_names, openpyxl_names)
        self.assertEqual(elixcee_names.get("test"), "Sheet1!$F$5")

    def test_defined_names_is_empty_on_a_fixture_with_none(self):
        vm = elixcee.load_workbook(FIXTURE)
        wb = openpyxl.load_workbook(FIXTURE)
        self.assertEqual(vm.defined_names(), {})
        self.assertEqual(dict(wb.defined_names), {})


class SheetStateAgreesWithOpenpyxl(unittest.TestCase):
    # No real fixture in this repo has a hidden/veryHidden sheet (see
    # docs/openpyxl-gap-audit.md), so the fixture here is built WITH
    # openpyxl itself rather than loaded from compat/oracle-excel-com --
    # this compares reads only, no elixcee save() round trip involved.
    def test_sheet_state_matches_openpyxl_on_an_openpyxl_authored_fixture(self):
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "states.xlsx")
            wb = openpyxl.Workbook()
            wb.active.title = "Visible"
            wb.create_sheet("Hidden").sheet_state = "hidden"
            wb.create_sheet("VeryHidden").sheet_state = "veryHidden"
            wb.save(path)

            vm = elixcee.load_workbook(path)
            wb2 = openpyxl.load_workbook(path)

            for name in ("Visible", "Hidden", "VeryHidden"):
                self.assertEqual(vm.sheet_state(name), wb2[name].sheet_state)

    def test_sheet_state_does_not_yet_survive_an_elixcee_save(self):
        # Disclosed known gap, not a hypothetical: elixcee has no writer
        # support for `state="..."` yet (no real fixture to validate the
        # writer shape against -- see ROADMAP.md's known gaps), so a loaded
        # hidden sheet currently reverts to visible on ANY save, even a
        # no-op one. Pinned here so a future writer fix is a deliberate,
        # visible change to this test, not a silent behavior shift.
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "states_src.xlsx")
            wb = openpyxl.Workbook()
            wb.active.title = "Visible"
            wb.create_sheet("Hidden").sheet_state = "hidden"
            wb.save(src)

            vm = elixcee.load_workbook(src)
            self.assertEqual(vm.sheet_state("Hidden"), "hidden")

            out = os.path.join(d, "states_out.xlsx")
            vm.save_workbook(out)
            wb2 = openpyxl.load_workbook(out)
            self.assertEqual(
                wb2["Hidden"].sheet_state,
                "visible",
                "known gap: update this test (and ship set_sheet_state) once "
                "the writer preserves state=... on save",
            )


class RowHeightAndColumnWidthAgreeWithOpenpyxl(unittest.TestCase):
    # No real fixture in this repo has a genuine custom row height or column
    # width (fixture1's only <col> is a hidden column with width="0", not
    # real data) -- fixture built with openpyxl itself, same approach as
    # SheetStateAgreesWithOpenpyxl. Reads only, no elixcee save() round trip.
    def test_row_height_and_column_width_match_openpyxl_on_an_openpyxl_authored_fixture(
        self,
    ):
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "dims.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["A1"] = 1
            ws.row_dimensions[5].height = 30.5
            ws.column_dimensions["B"].width = 12.5
            wb.save(path)

            vm = elixcee.load_workbook(path)
            wb2 = openpyxl.load_workbook(path)
            ws2 = wb2.active

            self.assertEqual(vm.row_height(5), ws2.row_dimensions[5].height)
            self.assertEqual(vm.column_width(2), ws2.column_dimensions["B"].width)
            # A row with no explicit height is None on both sides. Column is
            # NOT compared the same way: openpyxl's column_dimensions[letter]
            # auto-vivifies a default-width (13.0) entry on first `[]` access
            # even for a column the file never set -- an openpyxl
            # implementation artifact, not something elixcee's own correct
            # None-for-unset behavior should be judged against.
            self.assertIsNone(vm.row_height(1))
            self.assertIsNone(ws2.row_dimensions[1].height)
            self.assertIsNone(vm.column_width(1))

    def test_row_height_and_column_width_survive_an_elixcee_save(self):
        # Was a disclosed known gap (see this test's old name/history): the
        # writer used to unconditionally regenerate <row>/<cols> from
        # hidden-row/column state alone, dropping a loaded row height/column
        # width on EVERY save. Fixed -- see CHANGELOG.md. write support
        # (set_row_height/set_column_width) is still deferred separately (no
        # real fixture has genuine data to validate that API's own writer
        # shape against, a different concern from preserving an
        # already-loaded value through a save).
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "dims_src.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.row_dimensions[5].height = 30.5
            ws.column_dimensions["B"].width = 12.5
            wb.save(src)

            vm = elixcee.load_workbook(src)
            self.assertEqual(vm.row_height(5), 30.5)
            self.assertEqual(vm.column_width(2), 12.5)

            out = os.path.join(d, "dims_out.xlsx")
            vm.save_workbook(out)
            wb2 = openpyxl.load_workbook(out)
            ws2 = wb2.active
            self.assertEqual(ws2.row_dimensions[5].height, 30.5)
            self.assertEqual(ws2.column_dimensions["B"].width, 12.5)

            vm2 = elixcee.load_workbook(out)
            self.assertEqual(vm2.row_height(5), 30.5)
            self.assertEqual(vm2.column_width(2), 12.5)

            # A second save must not re-drop or duplicate the attributes.
            out2 = os.path.join(d, "dims_out2.xlsx")
            vm2.save_workbook(out2)
            vm3 = elixcee.load_workbook(out2)
            self.assertEqual(vm3.row_height(5), 30.5)
            self.assertEqual(vm3.column_width(2), 12.5)


class AutoFilterSurvivesAnElixceeSave(unittest.TestCase):
    # Was a disclosed known gap (ROADMAP.md known-gap 28, found while scoping
    # 0.14.0-C): `<autoFilter>` has no `r:id` at all, unlike tableParts/drawing, so
    # it needed no rels_survived gate -- but it was ALSO not one of the
    # unconditionally-restored opaque fragments (sheetFormatPr/dataValidations/
    # pageMargins/...), so it was silently destroyed on every elixcee save, not
    # merely stale. Fixed as a byte-preservation-only passthrough fragment -- see
    # CHANGELOG.md. elixcee has no auto_filter read/write API of its own (this is
    # preservation only, not new feature support -- create/remove/filter-type API
    # is 0.16.0), so this reads back through openpyxl on both sides rather than
    # comparing against an elixcee-side getter.
    def test_autofilter_ref_survives_an_elixcee_save(self):
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "af_src.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["A1"] = "h"
            ws.auto_filter.ref = "A1:C10"
            wb.save(src)

            vm = elixcee.load_workbook(src)
            out = os.path.join(d, "af_out.xlsx")
            vm.save_workbook(out)

            wb2 = openpyxl.load_workbook(out)
            self.assertEqual(wb2.active.auto_filter.ref, "A1:C10")

            # A second save must not re-drop or duplicate it.
            vm2 = elixcee.load_workbook(out)
            out2 = os.path.join(d, "af_out2.xlsx")
            vm2.save_workbook(out2)
            wb3 = openpyxl.load_workbook(out2)
            self.assertEqual(wb3.active.auto_filter.ref, "A1:C10")

    def test_autofilter_with_a_filter_column_survives_an_elixcee_save(self):
        # Child <filterColumn>/<filters>/<filter> content must round-trip too, not
        # just the container's own `ref` attribute -- extract_raw_element captures
        # the whole element span verbatim, so this also guards against a future
        # regression that tried to reconstruct the container from parsed parts.
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "af_cols_src.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.auto_filter.ref = "A1:C10"
            ws.auto_filter.add_filter_column(0, ["foo", "bar"], blank=False)
            wb.save(src)

            vm = elixcee.load_workbook(src)
            out = os.path.join(d, "af_cols_out.xlsx")
            vm.save_workbook(out)

            ws2 = openpyxl.load_workbook(out).active
            self.assertEqual(ws2.auto_filter.ref, "A1:C10")
            self.assertEqual(len(ws2.auto_filter.filterColumn), 1)
            col = ws2.auto_filter.filterColumn[0]
            self.assertEqual(col.colId, 0)
            self.assertEqual(sorted(col.filters.filter), ["bar", "foo"])

    def test_autofilter_survives_an_unrelated_cell_edit_and_a_real_merge(self):
        # Confirms schema position too: CT_Worksheet orders autoFilter BEFORE
        # mergeCells (verified against openpyxl's own writer, worksheet/_writer.py's
        # write_tail) -- an otherwise-correct-looking swap would produce a file
        # real Excel silently "repairs" by dropping the misordered element, even
        # though every byte is present.
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "af_merge_src.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["A1"] = "h"
            ws.merge_cells("D1:E1")
            ws.auto_filter.ref = "A1:C10"
            wb.save(src)

            vm = elixcee.load_workbook(src)
            vm.set_range("A2", [["unrelated edit"]])
            out = os.path.join(d, "af_merge_out.xlsx")
            vm.save_workbook(out)

            xml = zipfile.ZipFile(out).read("xl/worksheets/sheet1.xml").decode()
            af_pos = xml.find("<autoFilter")
            mc_pos = xml.find("<mergeCells")
            self.assertNotEqual(af_pos, -1)
            self.assertNotEqual(mc_pos, -1)
            self.assertLess(af_pos, mc_pos)

            ws2 = openpyxl.load_workbook(out).active
            self.assertEqual(ws2.auto_filter.ref, "A1:C10")
            self.assertEqual(ws2["A2"].value, "unrelated edit")
            self.assertIn("D1:E1", [str(r) for r in ws2.merged_cells.ranges])


class ConditionalFormattingSurvivesAnElixceeSave(unittest.TestCase):
    # Found while scoping 0.16.0 (Tables, Filters and Rules): <conditionalFormatting>
    # was silently dropped on EVERY elixcee save -- confirmed against a real fixture
    # (compat/oracle-excel-com/fixtures/pristine/fixture3_table_validation_conditional.xlsm)
    # before this fix. Same shape as the autoFilter fix above (a non-relationship-backed
    # element, unconditional opaque-fragment passthrough), except a sheet can carry more
    # than one <conditionalFormatting> block (one per range/rule-set), unlike autoFilter's
    # single element -- see reader::extract_all_raw_elements. `<dxfs>` (in xl/styles.xml,
    # which a rule's dxfId references) is never touched by any style-mutation resolve
    # pass, so a preserved rule's dxfId stays valid regardless. No create/edit API here
    # (that's real 0.16.0 feature work, a separate effort) -- preservation only.
    def test_a_single_rule_survives_an_elixcee_save(self):
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "cf_src.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["A1"] = 5
            rule = CellIsRule(
                operator="greaterThan",
                formula=["10"],
                font=openpyxl.styles.Font(color="9C0006"),
            )
            ws.conditional_formatting.add("A1:A5", rule)
            wb.save(src)

            vm = elixcee.load_workbook(src)
            out = os.path.join(d, "cf_out.xlsx")
            vm.save_workbook(out)

            ws2 = openpyxl.load_workbook(out).active
            rules = list(ws2.conditional_formatting)
            self.assertEqual(len(rules), 1)
            self.assertEqual(str(rules[0].sqref), "A1:A5")
            cf_rules = ws2.conditional_formatting[str(rules[0].sqref)]
            self.assertEqual(len(cf_rules), 1)
            self.assertEqual(cf_rules[0].operator, "greaterThan")
            self.assertEqual(cf_rules[0].formula, ["10"])

            # A second save must not re-drop or duplicate it.
            vm2 = elixcee.load_workbook(out)
            out2 = os.path.join(d, "cf_out2.xlsx")
            vm2.save_workbook(out2)
            ws3 = openpyxl.load_workbook(out2).active
            self.assertEqual(len(list(ws3.conditional_formatting)), 1)

    def test_multiple_rule_blocks_all_survive(self):
        # CT_Worksheet's conditionalFormatting is maxOccurs="unbounded" -- confirm more
        # than one range/rule-set in the same sheet round-trips, not just the common
        # single-block case every real fixture in this repo happens to have.
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "cf_multi_src.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["A1"] = 5
            ws["B1"] = -5
            ws.conditional_formatting.add(
                "A1:A5",
                CellIsRule(operator="greaterThan", formula=["10"], font=openpyxl.styles.Font(color="9C0006")),
            )
            ws.conditional_formatting.add(
                "B1:B5",
                CellIsRule(operator="lessThan", formula=["0"], font=openpyxl.styles.Font(color="006100")),
            )
            wb.save(src)

            vm = elixcee.load_workbook(src)
            out = os.path.join(d, "cf_multi_out.xlsx")
            vm.save_workbook(out)

            ws2 = openpyxl.load_workbook(out).active
            sqrefs = {str(r.sqref) for r in ws2.conditional_formatting}
            self.assertEqual(sqrefs, {"A1:A5", "B1:B5"})

    def test_survives_an_unrelated_style_edit_and_confirms_schema_position(self):
        # Confirms schema position (verified against fixture3/fixture4's real bytes:
        # phoneticPr -> conditionalFormatting -> dataValidations) and that set_style
        # (which mutates <cellXfs>, never <dxfs>) doesn't disturb a preserved rule's
        # dxfId reference.
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "cf_style_src.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["A1"] = 5
            ws["C1"] = "unrelated"
            ws.conditional_formatting.add(
                "A1:A5",
                CellIsRule(operator="greaterThan", formula=["10"], font=openpyxl.styles.Font(color="9C0006")),
            )
            wb.save(src)

            vm = elixcee.load_workbook(src)
            vm.set_style("C1", font={"bold": True})
            out = os.path.join(d, "cf_style_out.xlsx")
            vm.save_workbook(out)

            xml = zipfile.ZipFile(out).read("xl/worksheets/sheet1.xml").decode()
            cf_pos = xml.find("<conditionalFormatting")
            dv_pos = xml.find("<dataValidations")
            self.assertNotEqual(cf_pos, -1)
            if dv_pos != -1:
                self.assertLess(cf_pos, dv_pos)

            ws2 = openpyxl.load_workbook(out).active
            rules = list(ws2.conditional_formatting)
            self.assertEqual(len(rules), 1)
            cf_rules = ws2.conditional_formatting[str(rules[0].sqref)]
            self.assertEqual(cf_rules[0].operator, "greaterThan")
            self.assertTrue(ws2["C1"].font.bold)


class FromScratchVmProducesAnOpenpyxlReadableStylesheet(unittest.TestCase):
    # A bare elixcee.Vm() (no loaded source file) used to emit a minimal
    # styles.xml with bare <fill/> elements -- no <patternFill>/<gradientFill>
    # child -- which openpyxl's own reader rejects with
    # TypeError: expected <class 'openpyxl.styles.fills.Fill'>. Fixed by
    # matching openpyxl's own from-scratch default shape:
    # <fill><patternFill/></fill> (index 0) and
    # <fill><patternFill patternType="gray125"/></fill> (index 1).
    def test_from_scratch_vm_save_reopens_cleanly_in_openpyxl(self):
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "from_scratch.xlsx")
            vm = elixcee.Vm()
            vm.set_cell(1, 1, "hello")
            vm.save_workbook(out)

            xml = zipfile.ZipFile(out).read("xl/styles.xml").decode()
            self.assertIn('<fill><patternFill/></fill>', xml)
            self.assertIn(
                '<fill><patternFill patternType="gray125"/></fill>', xml
            )

            ws = openpyxl.load_workbook(out).active
            self.assertEqual(ws["A1"].value, "hello")

            # A second save/reload cycle (openpyxl re-saving what it just
            # read) must also round-trip cleanly -- confirms the shape isn't
            # just tolerated once but genuinely well-formed.
            out2 = os.path.join(d, "from_scratch_resaved.xlsx")
            ws.parent.save(out2)
            ws2 = openpyxl.load_workbook(out2).active
            self.assertEqual(ws2["A1"].value, "hello")


class SetStyleAgreesWithOpenpyxl(unittest.TestCase):
    # 0.15.0-B: `set_style(font=..., fill=..., border=..., alignment=..., protection=...)`.
    # xl/styles.xml was 100% opaque passthrough until 0.15.0-A/B; <fonts>/<fills>/
    # <borders> are now found-or-appended the same way <cellXfs>/<numFmts> already
    # were. Fill/border/protection/most-alignment properties have NO real Excel
    # fixture in this repo (ROADMAP.md's 0.15.0-B entry) -- user-granted exception to
    # this project's usual "no writer without a real fixture" gate, same basis as
    # 0.15.0-A's custom-numFmt path: unambiguous ECMA-376 spec shape, no producer
    # variance. So every case here is verified via openpyxl (a second, independent
    # OOXML consumer) rather than a real-Excel-authored source file.
    def test_font_bold_survives_a_save_and_reopens_clean(self):
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "font.xlsx")
            vm = elixcee.Vm()
            vm.set_cell(1, 1, "hi")
            vm.set_style("A1", font={"bold": True})
            vm.save_workbook(out)
            ws = openpyxl.load_workbook(out).active
            self.assertTrue(ws["A1"].font.bold)

    def test_font_edit_preserves_a_real_theme_colored_underlined_font(self):
        # fixture4's real, in-use hyperlink font: underlined, theme-colored, sized --
        # cloning it while only setting `bold` must not lose the other properties.
        fixture = os.path.join(
            os.path.dirname(__file__),
            "..",
            "oracle-excel-com",
            "fixtures",
            "pristine",
            "fixture4_hyperlink_comment_name.xlsm",
        )
        with tempfile.TemporaryDirectory() as d:
            vm = elixcee.load_workbook(fixture)
            out = os.path.join(d, "font_theme.xlsm")
            vm.set_style("A1", font={"bold": True})
            vm.save_workbook(out)
            xml = zipfile.ZipFile(out).read("xl/styles.xml").decode()
            self.assertIn('theme="10"', xml)

    def test_solid_fill_uses_fgcolor_and_indexed_64_bgcolor(self):
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "fill.xlsx")
            vm = elixcee.Vm()
            vm.set_cell(1, 1, "hi")
            vm.set_style("A1", fill={"type": "solid", "color": "4472C4"})
            vm.save_workbook(out)
            ws = openpyxl.load_workbook(out).active
            self.assertEqual(ws["A1"].fill.fgColor.rgb, "FF4472C4")

    def test_border_touches_only_the_requested_side(self):
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "border.xlsx")
            vm = elixcee.Vm()
            vm.set_cell(1, 1, "hi")
            vm.set_style("A1", border={"top": {"style": "thick", "color": "FF000000"}})
            vm.save_workbook(out)
            ws = openpyxl.load_workbook(out).active
            self.assertEqual(ws["A1"].border.top.style, "thick")
            left = ws["A1"].border.left
            self.assertTrue(left is None or left.style is None)

    def test_alignment_merges_onto_an_existing_attribute_not_replace(self):
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "align.xlsx")
            vm = elixcee.Vm()
            vm.set_cell(1, 1, "hi")
            vm.set_style("A1", alignment={"vertical": "center"})
            vm.set_style("A1", alignment={"horizontal": "center"})
            vm.save_workbook(out)
            ws = openpyxl.load_workbook(out).active
            self.assertEqual(ws["A1"].alignment.vertical, "center")
            self.assertEqual(ws["A1"].alignment.horizontal, "center")

    def test_protection_locked_and_hidden(self):
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "protect.xlsx")
            vm = elixcee.Vm()
            vm.set_cell(1, 1, "hi")
            vm.set_style("A1", protection={"locked": False, "hidden": True})
            vm.save_workbook(out)
            ws = openpyxl.load_workbook(out).active
            self.assertFalse(ws["A1"].protection.locked)
            self.assertTrue(ws["A1"].protection.hidden)

    def test_two_calls_on_the_same_cell_accumulate_instead_of_overwriting(self):
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "accum.xlsx")
            vm = elixcee.Vm()
            vm.set_cell(1, 1, "hi")
            vm.set_style("A1", font={"bold": True})
            vm.set_style("A1", fill={"type": "solid", "color": "00FF00"})
            vm.save_workbook(out)
            ws = openpyxl.load_workbook(out).active
            self.assertTrue(ws["A1"].font.bold)
            self.assertEqual(ws["A1"].fill.fgColor.rgb, "FF00FF00")

    def test_set_number_format_and_set_style_on_the_same_cell_both_survive(self):
        # The chaining fix: 0.15.0-A's own resolve pass must not be silently
        # overwritten by 0.15.0-B's, or vice versa, when both touch one cell before
        # a single save.
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "chain.xlsx")
            vm = elixcee.Vm()
            vm.set_cell(1, 1, 1234.5)
            vm.set_number_format("A1", "#,##0.00")
            vm.set_style("A1", font={"bold": True})
            vm.save_workbook(out)
            ws = openpyxl.load_workbook(out).active
            self.assertEqual(ws["A1"].number_format, "#,##0.00")
            self.assertTrue(ws["A1"].font.bold)

    def test_does_not_mutate_a_style_shared_with_an_untouched_cell(self):
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "safety.xlsx")
            vm = elixcee.Vm()
            vm.set_cell(1, 1, "a")
            vm.set_cell(2, 1, "b")
            vm.set_style("A1", font={"bold": True})
            vm.save_workbook(out)
            ws = openpyxl.load_workbook(out).active
            self.assertTrue(ws["A1"].font.bold)
            self.assertIsNot(ws["A2"].font.bold, True)

    def test_a_second_save_reload_cycle_is_stable(self):
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "once.xlsx")
            vm = elixcee.Vm()
            vm.set_cell(1, 1, "hi")
            vm.set_style("A1", font={"bold": True}, fill={"type": "solid", "color": "00FF00"})
            vm.save_workbook(out)

            out2 = os.path.join(d, "twice.xlsx")
            vm2 = elixcee.load_workbook(out)
            vm2.save_workbook(out2)
            ws = openpyxl.load_workbook(out2).active
            self.assertTrue(ws["A1"].font.bold)
            self.assertEqual(ws["A1"].fill.fgColor.rgb, "FF00FF00")


class TablesAgreeWithOpenpyxl(unittest.TestCase):
    # 0.16.0-A1: read-only tables() against fixture3's real, complete table
    # (name/displayName "テーブル1", 3 columns, TableStyleMedium2, a nested bare
    # <autoFilter>). No create/edit/delete API yet -- that's 0.16.0-A2/A3.
    def test_tables_reports_fixture3s_real_table_data(self):
        vm = elixcee.load_workbook(FIXTURE3)
        tables = vm.tables("Sheet1")
        self.assertEqual(len(tables), 1)
        t = tables[0]
        self.assertEqual(t["name"], "テーブル1")
        self.assertEqual(t["display_name"], "テーブル1")
        self.assertEqual(t["ref"], "A1:C4")
        self.assertEqual(t["header_row_count"], 1)
        self.assertEqual(t["totals_row_count"], 0)
        self.assertFalse(t["totals_row_shown"])
        self.assertEqual(t["style_name"], "TableStyleMedium2")
        self.assertEqual(t["auto_filter_ref"], "A1:C4")
        self.assertEqual([c["name"] for c in t["columns"]], ["Name", "Qty", "Status"])
        self.assertTrue(all(c["calculated_column_formula"] is None for c in t["columns"]))

    def test_an_unmodified_table_survives_an_unrelated_save_byte_identical(self):
        with zipfile.ZipFile(FIXTURE3) as z:
            source_table_xml = z.read("xl/tables/table1.xml")

        vm = elixcee.load_workbook(FIXTURE3)
        vm.set_cell(10, 10, 999)  # unrelated, nowhere near the table
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "tables_out.xlsx")
            vm.save_workbook(out)
            with zipfile.ZipFile(out) as z:
                out_table_xml = z.read("xl/tables/table1.xml")
            self.assertEqual(out_table_xml, source_table_xml)

            ws2 = openpyxl.load_workbook(out)["Sheet1"]
            self.assertIn("テーブル1", ws2.tables)

            # Second save-reload cycle must not drift.
            vm2 = elixcee.load_workbook(out)
            out2 = os.path.join(d, "tables_out2.xlsx")
            vm2.save_workbook(out2)
            with zipfile.ZipFile(out2) as z:
                self.assertEqual(z.read("xl/tables/table1.xml"), source_table_xml)

    def test_a_row_insert_above_the_table_shifts_its_reported_ref_and_auto_filter(self):
        vm = elixcee.load_workbook(FIXTURE3)
        vm.insert_rows(1, amount=2, sheet="Sheet1")
        t = vm.tables("Sheet1")[0]
        self.assertEqual(t["ref"], "A3:C6")
        # Regression: the nested <autoFilter> covers the same area as the table's own
        # ref in every real fixture -- it must shift identically, not go stale.
        self.assertEqual(t["auto_filter_ref"], "A3:C6")

    def test_a_column_insert_before_the_table_shifts_its_reported_ref(self):
        vm = elixcee.load_workbook(FIXTURE3)
        vm.insert_cols(1, amount=1, sheet="Sheet1")
        t = vm.tables("Sheet1")[0]
        self.assertEqual(t["ref"], "B1:D4")

    # 0.16.0-A2: edit_table -- rename/resize/restyle/totals-row/column add-remove.
    def test_edit_table_rename_preserves_id_and_survives_openpyxl_reopen(self):
        with zipfile.ZipFile(FIXTURE3) as z:
            source_table_xml = z.read("xl/tables/table1.xml").decode("utf-8")

        vm = elixcee.load_workbook(FIXTURE3)
        vm.edit_table("テーブル1", display_name="Renamed")
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "out.xlsx")
            vm.save_workbook(out)
            t = elixcee.load_workbook(out).tables("Sheet1")[0]
            self.assertEqual(t["display_name"], "Renamed")
            with zipfile.ZipFile(out) as z:
                out_xml = z.read("xl/tables/table1.xml").decode("utf-8")
            # id/xr:uid untouched -- surgical patch, not a reserialize.
            self.assertIn('id="1"', out_xml)
            for tok in ("xr:uid=", "xr3:uid="):
                if tok in source_table_xml:
                    self.assertEqual(
                        source_table_xml.count(tok), out_xml.count(tok)
                    )
            openpyxl.load_workbook(out)  # must not raise

    def test_edit_table_resize_only_touches_the_tables_own_ref(self):
        vm = elixcee.load_workbook(FIXTURE3)
        vm.edit_table("テーブル1", ref="A1:C5")
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "out.xlsx")
            vm.save_workbook(out)
            t = elixcee.load_workbook(out).tables("Sheet1")[0]
            self.assertEqual(t["ref"], "A1:C5")
            self.assertEqual(t["auto_filter_ref"], "A1:C4")  # unaffected
            openpyxl.load_workbook(out)

    def test_edit_table_style_and_totals_row_toggle(self):
        vm = elixcee.load_workbook(FIXTURE3)
        vm.edit_table("テーブル1", style_name="TableStyleLight8", totals_row_shown=True)
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "out.xlsx")
            vm.save_workbook(out)
            t = elixcee.load_workbook(out).tables("Sheet1")[0]
            self.assertEqual(t["style_name"], "TableStyleLight8")
            self.assertTrue(t["totals_row_shown"])
            openpyxl.load_workbook(out)

    def test_edit_table_add_column_appends_at_the_right_edge_and_widens_ref(self):
        vm = elixcee.load_workbook(FIXTURE3)
        vm.edit_table("テーブル1", add_columns=["Total"])
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "out.xlsx")
            vm.save_workbook(out)
            t = elixcee.load_workbook(out).tables("Sheet1")[0]
            self.assertEqual(
                [c["name"] for c in t["columns"]], ["Name", "Qty", "Status", "Total"]
            )
            self.assertEqual(t["ref"], "A1:D4")  # widened, not left stale
            openpyxl.load_workbook(out)

    def test_edit_table_remove_column_deletes_and_shifts_cell_data_left(self):
        vm = elixcee.load_workbook(FIXTURE3)
        before_status = [vm.get_range(f"C{r}:C{r}", sheet="Sheet1")[0][0] for r in range(1, 5)]
        vm.edit_table("テーブル1", remove_columns=["Qty"])
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "out.xlsx")
            vm.save_workbook(out)
            vm2 = elixcee.load_workbook(out)
            t = vm2.tables("Sheet1")[0]
            self.assertEqual([c["name"] for c in t["columns"]], ["Name", "Status"])
            self.assertEqual(t["ref"], "A1:B4")  # narrowed, not left stale
            after_b = [vm2.get_range(f"B{r}:B{r}", sheet="Sheet1")[0][0] for r in range(1, 5)]
            self.assertEqual(after_b, before_status)  # Status shifted from C into B
            after_c = [vm2.get_range(f"C{r}:C{r}", sheet="Sheet1")[0][0] for r in range(1, 5)]
            self.assertTrue(all(v is None for v in after_c))  # vacated, not duplicated
            with warnings.catch_warnings(record=True) as w:
                warnings.simplefilter("always")
                openpyxl.load_workbook(out)
                self.assertEqual(w, [])

    def test_edit_table_rejects_unknown_table_or_column_name(self):
        vm = elixcee.load_workbook(FIXTURE3)
        with self.assertRaises(ValueError):
            vm.edit_table("NoSuchTable")
        with self.assertRaises(ValueError):
            vm.edit_table("テーブル1", remove_columns=["NoSuchColumn"])

    def test_a_structural_edit_shift_is_now_persisted_to_the_saved_file(self):
        # Regression: 0.16.0-A1's shift updated tables()'s report but never reached the
        # saved xl/tables/tableN.xml -- 0.16.0-A2 closes this as part of building the
        # write path 0.16.0-A1 deliberately left untouched.
        vm = elixcee.load_workbook(FIXTURE3)
        vm.insert_rows(1, amount=2, sheet="Sheet1")
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "out.xlsx")
            vm.save_workbook(out)
            with zipfile.ZipFile(out) as z:
                out_xml = z.read("xl/tables/table1.xml").decode("utf-8")
            self.assertIn('ref="A3:C6"', out_xml)
            t = elixcee.load_workbook(out).tables("Sheet1")[0]
            self.assertEqual(t["ref"], "A3:C6")
            self.assertEqual(t["auto_filter_ref"], "A3:C6")
            openpyxl.load_workbook(out)

    # 0.16.0-A3: create_table -- the three-part linkage (worksheet .rels, <tableParts>,
    # [Content_Types].xml) built from nothing, verified for both a from-scratch Vm() and
    # a loaded file that already has a table (the merge path, not the synthesis path).
    def test_create_table_from_scratch_reopens_cleanly_with_no_warnings(self):
        vm = elixcee.Vm()
        vm.set_cell(1, 1, "Name")
        vm.set_cell(1, 2, "Qty")
        vm.set_cell(2, 1, "Widget")
        vm.set_cell(2, 2, 5)
        vm.create_table("A1:B2", name="Table1")
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "out.xlsx")
            vm.save_workbook(out)
            with warnings.catch_warnings(record=True) as w:
                warnings.simplefilter("always")
                wb = openpyxl.load_workbook(out)
                self.assertEqual(w, [])
            ws = wb.active
            self.assertIn("Table1", ws.tables)
            t = ws.tables["Table1"]
            self.assertEqual(t.ref, "A1:B2")
            self.assertEqual([c.name for c in t.tableColumns], ["Name", "Qty"])

    def test_create_table_alongside_an_existing_table_preserves_it_byte_identical(self):
        with zipfile.ZipFile(FIXTURE3) as z:
            source_table1_xml = z.read("xl/tables/table1.xml")

        vm = elixcee.load_workbook(FIXTURE3)
        vm.set_cell(1, 5, "City")
        vm.set_cell(1, 6, "Pop")
        vm.create_table("E1:F1", name="Table2", sheet="Sheet1")
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "out.xlsm")
            vm.save_workbook(out)
            with zipfile.ZipFile(out) as z:
                self.assertEqual(z.read("xl/tables/table1.xml"), source_table1_xml)
                rels = z.read("xl/worksheets/_rels/sheet1.xml.rels").decode("utf-8")
                self.assertIn("table1.xml", rels)
                self.assertIn("table2.xml", rels)
                ct = z.read("[Content_Types].xml").decode("utf-8")
                self.assertIn("/xl/tables/table1.xml", ct)
                self.assertIn("/xl/tables/table2.xml", ct)

            with warnings.catch_warnings(record=True) as w:
                warnings.simplefilter("always")
                wb = openpyxl.load_workbook(out, keep_vba=True)
                self.assertEqual(w, [])
            ws = wb["Sheet1"]
            self.assertIn("テーブル1", ws.tables)
            self.assertIn("Table2", ws.tables)

            # Second save-reload cycle must not drift or duplicate anything.
            vm2 = elixcee.load_workbook(out)
            out2 = os.path.join(d, "out2.xlsm")
            vm2.save_workbook(out2)
            with zipfile.ZipFile(out) as z1, zipfile.ZipFile(out2) as z2:
                for name in ("xl/tables/table1.xml", "xl/tables/table2.xml",
                             "xl/worksheets/_rels/sheet1.xml.rels"):
                    self.assertEqual(z1.read(name), z2.read(name))

    def test_create_table_rejects_an_overlapping_range(self):
        vm = elixcee.load_workbook(FIXTURE3)  # already has テーブル1 at A1:C4
        with self.assertRaises(ValueError):
            vm.create_table("B2:D5", name="Overlap")

    def test_create_table_rejects_a_blank_header_cell(self):
        vm = elixcee.load_workbook(FIXTURE3)
        vm.set_cell(1, 10, "OnlyOne")  # J1 set, K1 deliberately blank
        with self.assertRaises(ValueError):
            vm.create_table("J1:K1", name="Blank")

    def test_create_table_requires_name_or_display_name(self):
        vm = elixcee.Vm()
        vm.set_cell(1, 1, "Col")
        with self.assertRaises(ValueError):
            vm.create_table("A1:A1")


class DataValidationAgreesWithOpenpyxl(unittest.TestCase):
    # 0.16.0-C: add_data_validation()/remove_data_validation()/data_validations()
    # against fixture3's real, complete <dataValidation type="list" sqref="E1"
    # xr:uid="{...}">, plus openpyxl-authored synthetic fixtures for the 5
    # validation types/multi-area sqref this project has no real-Excel example of
    # (granted verification exception, same basis as every prior style-engine/table
    # exception this session -- see internal_docs/data-validation-0.16.0-c-design.md).
    def test_data_validations_reports_fixture3s_real_rule(self):
        vm = elixcee.load_workbook(FIXTURE3)
        rules = vm.data_validations("Sheet1")
        self.assertEqual(len(rules), 1)
        r = rules[0]
        self.assertEqual(r["validation_type"], "list")
        self.assertEqual(r["sqref"], ["E1"])
        self.assertEqual(r["formula1"], '"Yes,No,Maybe"')

    def test_an_unmodified_rule_survives_an_unrelated_save_byte_identical(self):
        xml_before = zipfile.ZipFile(FIXTURE3).read("xl/worksheets/sheet1.xml").decode()
        vm = elixcee.load_workbook(FIXTURE3)
        vm.set_cell(10, 10, 999)  # unrelated, nowhere near the rule
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "dv_out.xlsx")
            vm.save_workbook(out)
            xml_after = zipfile.ZipFile(out).read("xl/worksheets/sheet1.xml").decode()

            def dv_span(xml):
                start = xml.index("<dataValidations")
                end = xml.index("</dataValidations>") + len("</dataValidations>")
                return xml[start:end]

            self.assertEqual(dv_span(xml_before), dv_span(xml_after))

    def test_add_a_list_rule_and_reopen_with_openpyxl(self):
        vm = elixcee.load_workbook(FIXTURE3)
        idx = vm.add_data_validation("B1:B5", "list", formula1='"A,B,C"', sheet="Sheet1")
        self.assertEqual(idx, 1)  # fixture3 already has one real rule (index 0)
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "dv_add.xlsx")
            vm.save_workbook(out)
            rules = openpyxl.load_workbook(out).active.data_validations.dataValidation
            self.assertEqual(len(rules), 2)
            self.assertEqual(sorted(r.type for r in rules), ["list", "list"])
            self.assertEqual(sorted(str(r.sqref) for r in rules), ["B1:B5", "E1"])

    def test_remove_the_only_rule_omits_the_container_entirely(self):
        vm = elixcee.load_workbook(FIXTURE3)
        vm.remove_data_validation(0, sheet="Sheet1")
        self.assertEqual(vm.data_validations("Sheet1"), [])
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "dv_removed.xlsx")
            vm.save_workbook(out)
            xml = zipfile.ZipFile(out).read("xl/worksheets/sheet1.xml").decode()
            self.assertNotIn("dataValidations", xml)
            # Must not raise/warn on reopen -- an empty <dataValidations/> would be
            # invalid OOXML (CT_DataValidations' own child is minOccurs="1").
            openpyxl.load_workbook(out)

    def test_every_validation_type_round_trips_through_a_synthetic_fixture(self):
        cases = [
            ("whole", dict(type="whole", operator="between", formula1="1", formula2="10")),
            ("decimal", dict(type="decimal", operator="greaterThan", formula1="0.5")),
            ("date", dict(type="date", operator="lessThan", formula1="2025-01-01")),
            ("time", dict(type="time", operator="lessThanOrEqual", formula1="0.5")),
            ("textLength", dict(type="textLength", operator="equal", formula1="5")),
            ("custom", dict(type="custom", formula1="ISNUMBER(A1)")),
        ]
        for name, kwargs in cases:
            with self.subTest(validation_type=name), tempfile.TemporaryDirectory() as d:
                from openpyxl.worksheet.datavalidation import DataValidation

                wb = openpyxl.Workbook()
                ws = wb.active
                dv = DataValidation(**kwargs)
                dv.add("A1:A2")
                ws.add_data_validation(dv)
                ws["A1"] = 1
                src = os.path.join(d, f"dv_{name}_src.xlsx")
                wb.save(src)

                vm = elixcee.load_workbook(src)
                rules = vm.data_validations()
                self.assertEqual(len(rules), 1)
                self.assertEqual(rules[0]["validation_type"], kwargs["type"])

                out = os.path.join(d, f"dv_{name}_out.xlsx")
                vm.save_workbook(out)
                reopened = openpyxl.load_workbook(out).active.data_validations.dataValidation
                self.assertEqual(len(reopened), 1)
                self.assertEqual(reopened[0].type, kwargs["type"])

    def test_multi_area_sqref_round_trips(self):
        vm = elixcee.Vm()
        vm.set_cell(1, 1, 1)
        vm.add_data_validation("A1:A5 C1:C5", "whole", operator="greaterThan", formula1="0")
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "dv_multiarea.xlsx")
            vm.save_workbook(out)
            rule = openpyxl.load_workbook(out).active.data_validations.dataValidation[0]
            self.assertIn(str(rule.sqref), ("A1:A5 C1:C5", "C1:C5 A1:A5"))

    def test_a_structural_edit_shift_is_persisted_to_disk_not_just_in_memory(self):
        # The critical regression: shift_data_validations_for_structural_edit only
        # updates in-memory `sqref` immediately -- the ON-DISK bytes only reflect it
        # once `dirty` routes through `resolve_data_validations_for_sheet` at save
        # time. A bug here would leave the saved file stale while the in-memory read
        # API reports the correct (but never-persisted) shifted range.
        vm = elixcee.load_workbook(FIXTURE3)
        vm.insert_rows(1, amount=1, sheet="Sheet1")
        self.assertEqual(vm.data_validations("Sheet1")[0]["sqref"], ["E2"])
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "dv_shift.xlsx")
            vm.save_workbook(out)
            xml = zipfile.ZipFile(out).read("xl/worksheets/sheet1.xml").decode()
            self.assertIn('sqref="E2"', xml)
            self.assertNotIn('sqref="E1"', xml)
            rule = openpyxl.load_workbook(out).active.data_validations.dataValidation[0]
            self.assertEqual(str(rule.sqref), "E2")

            # Second save-reload cycle must not drift or duplicate.
            vm2 = elixcee.load_workbook(out)
            out2 = os.path.join(d, "dv_shift2.xlsx")
            vm2.save_workbook(out2)
            rules2 = openpyxl.load_workbook(out2).active.data_validations.dataValidation
            self.assertEqual(len(rules2), 1)
            self.assertEqual(str(rules2[0].sqref), "E2")


class AutoFilterFilteringAgreesWithOpenpyxl(unittest.TestCase):
    # 0.16.0-B1: add_autofilter()/set_equality_filter()/set_custom_filter()/
    # set_blank_filter()/set_top10_filter()/set_date_group_filter()/
    # clear_filter_column()/remove_autofilter()/autofilter() -- distinct from
    # AutoFilterSurvivesAnElixceeSave above (PR #29's byte-preservation fix for an
    # ALREADY-PRESENT `<autoFilter>`); this class covers actually authoring filter
    # criteria and the real row-hide evaluation they trigger. No real fixture has an
    # active filter criterion of any kind (verified during scoping) -- every case here
    # uses an openpyxl-authored synthetic fixture, the granted verification exception
    # (see internal_docs/autofilter-0.16.0-b-design.md).

    def test_equality_filter_hides_non_matching_rows_and_round_trips(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name"])
        for v in ["A", "B", "C", "A", "B"]:
            ws.append([v])
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "af_eq_src.xlsx")
            wb.save(src)

            vm = elixcee.load_workbook(src)
            vm.add_autofilter("A1:A6")
            vm.set_equality_filter(0, ["A", "B"])
            self.assertEqual(vm.hidden_rows(), [4])  # row 4 = "C"

            out = os.path.join(d, "af_eq_out.xlsx")
            vm.save_workbook(out)
            reopened = openpyxl.load_workbook(out)
            self.assertIsNotNone(reopened)
            xml = zipfile.ZipFile(out).read("xl/worksheets/sheet1.xml").decode()
            self.assertIn("<filterColumn", xml)
            self.assertIn("<filters>", xml)

            # Round trip + second save cycle: criteria and hidden state both stable.
            vm2 = elixcee.load_workbook(out)
            self.assertEqual(vm2.autofilter()["columns"][0]["values"], ["A", "B"])
            self.assertEqual(vm2.hidden_rows(), [4])
            out2 = os.path.join(d, "af_eq_out2.xlsx")
            vm2.save_workbook(out2)
            vm3 = elixcee.load_workbook(out2)
            self.assertEqual(vm3.autofilter()["columns"][0]["values"], ["A", "B"])

    def test_custom_filter_loads_real_openpyxl_criteria_and_combines_two_conditions_via_and(
        self,
    ):
        from openpyxl.worksheet.filters import AutoFilter, CustomFilter, CustomFilters, FilterColumn

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Num"])
        for v in [1, 5, 10, 15, 20]:
            ws.append([v])
        af = AutoFilter(ref="A1:A6")
        af.filterColumn.append(
            FilterColumn(
                colId=0,
                customFilters=CustomFilters(
                    customFilter=[CustomFilter(operator="greaterThan", val="5")], _and=False
                ),
            )
        )
        ws.auto_filter = af
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "af_custom_src.xlsx")
            wb.save(src)

            vm = elixcee.load_workbook(src)
            loaded = vm.autofilter()
            self.assertEqual(loaded["columns"][0]["operator"], "greaterThan")
            # Loading never auto-evaluates -- openpyxl itself never computed/wrote any
            # <row hidden="1"> either, matching this milestone's "evaluate once, on an
            # explicit call" design, not automatically on load.
            self.assertEqual(vm.hidden_rows(), [])

            vm.set_custom_filter(
                0, "greaterThanOrEqual", "10", and_=True, operator2="lessThanOrEqual", value2="15"
            )
            self.assertEqual(vm.hidden_rows(), [2, 3, 6])  # 1, 5, 20 fail; 10, 15 pass

            out = os.path.join(d, "af_custom_out.xlsx")
            vm.save_workbook(out)
            self.assertIsNotNone(openpyxl.load_workbook(out))
            xml = zipfile.ZipFile(out).read("xl/worksheets/sheet1.xml").decode()
            self.assertIn('and="1"', xml)
            self.assertIn('<customFilter operator="greaterThanOrEqual" val="10"', xml)
            self.assertIn('<customFilter operator="lessThanOrEqual" val="15"', xml)

    def test_blank_filter_hides_non_blank_rows(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["X"])
        ws.append(["a"])
        ws.append([None])
        ws.append(["b"])
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "af_blank_src.xlsx")
            wb.save(src)
            vm = elixcee.load_workbook(src)
            vm.add_autofilter("A1:A4")
            vm.set_blank_filter(0)
            self.assertEqual(vm.hidden_rows(), [2, 4])
            out = os.path.join(d, "af_blank_out.xlsx")
            vm.save_workbook(out)
            self.assertIsNotNone(openpyxl.load_workbook(out))
            xml = zipfile.ZipFile(out).read("xl/worksheets/sheet1.xml").decode()
            self.assertIn('<filters blank="1"', xml)

    def test_top10_filter_keeps_the_highest_n_real_values(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["N"])
        for v in [10, 20, 30, 40, 50]:
            ws.append([v])
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "af_top10_src.xlsx")
            wb.save(src)
            vm = elixcee.load_workbook(src)
            vm.add_autofilter("A1:A6")
            vm.set_top10_filter(0, 2, top=True, percent=False)
            self.assertEqual(vm.hidden_rows(), [2, 3, 4])  # only 40, 50 (rows 5, 6) stay
            out = os.path.join(d, "af_top10_out.xlsx")
            vm.save_workbook(out)
            self.assertIsNotNone(openpyxl.load_workbook(out))
            xml = zipfile.ZipFile(out).read("xl/worksheets/sheet1.xml").decode()
            self.assertIn("<top10", xml)

    def test_date_group_filter_matches_only_the_specified_month(self):
        import datetime

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["D"])
        ws.append([datetime.date(2024, 1, 15)])
        ws.append([datetime.date(2024, 2, 15)])
        ws.append([datetime.date(2024, 1, 20)])
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "af_dategroup_src.xlsx")
            wb.save(src)
            vm = elixcee.load_workbook(src)
            vm.add_autofilter("A1:A4")
            vm.set_date_group_filter(0, year=2024, month=1, grouping="month")
            self.assertEqual(vm.hidden_rows(), [3])  # the February row
            out = os.path.join(d, "af_dategroup_out.xlsx")
            vm.save_workbook(out)
            self.assertIsNotNone(openpyxl.load_workbook(out))
            xml = zipfile.ZipFile(out).read("xl/worksheets/sheet1.xml").decode()
            self.assertIn("<dateGroupItem", xml)

    def test_an_unmodified_autofilter_criterion_survives_an_unrelated_save_byte_identical(
        self,
    ):
        from openpyxl.worksheet.filters import AutoFilter, FilterColumn, Filters

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["X"])
        ws.append(["a"])
        af = AutoFilter(ref="A1:A2")
        af.filterColumn.append(FilterColumn(colId=0, filters=Filters(filter=["a"])))
        ws.auto_filter = af
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "af_preserve_src.xlsx")
            wb.save(src)
            xml_before = zipfile.ZipFile(src).read("xl/worksheets/sheet1.xml").decode()
            af_before = xml_before[xml_before.index("<autoFilter") : xml_before.index("</autoFilter>") + 13]

            vm = elixcee.load_workbook(src)
            vm.set_cell(1, 2, "unrelated edit")  # B1, nowhere near the filter
            out = os.path.join(d, "af_preserve_out.xlsx")
            vm.save_workbook(out)
            xml_after = zipfile.ZipFile(out).read("xl/worksheets/sheet1.xml").decode()
            af_after = xml_after[xml_after.index("<autoFilter") : xml_after.index("</autoFilter>") + 13]
            self.assertEqual(af_before, af_after)

    def test_remove_autofilter_reveals_every_hidden_row(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Name"])
        for v in ["A", "B", "C"]:
            ws.append([v])
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "af_remove_src.xlsx")
            wb.save(src)
            vm = elixcee.load_workbook(src)
            vm.add_autofilter("A1:A4")
            vm.set_equality_filter(0, ["A"])
            self.assertTrue(len(vm.hidden_rows()) > 0)
            vm.remove_autofilter()
            self.assertEqual(vm.hidden_rows(), [])
            self.assertIsNone(vm.autofilter())

    def test_clear_filter_column_leaves_other_columns_active(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["A", "B"])
        ws.append(["x", 10])
        ws.append(["y", 10])
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "af_clear_src.xlsx")
            wb.save(src)
            vm = elixcee.load_workbook(src)
            vm.add_autofilter("A1:B3")
            vm.set_equality_filter(0, ["x"])
            vm.clear_filter_column(0)
            self.assertEqual(vm.autofilter()["columns"], [])
            self.assertEqual(vm.hidden_rows(), [])

    def test_a_structural_edit_shift_is_persisted_to_disk_not_just_in_memory(self):
        from openpyxl.worksheet.filters import AutoFilter, FilterColumn, Filters

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["X"])
        ws.append(["a"])
        af = AutoFilter(ref="A1:A2")
        af.filterColumn.append(FilterColumn(colId=0, filters=Filters(filter=["a"])))
        ws.auto_filter = af
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "af_shift_src.xlsx")
            wb.save(src)
            vm = elixcee.load_workbook(src)
            # Inserting AT row 1 (the header's own row) shifts BOTH corners down, same
            # real Excel semantics as any other range whose top edge is at-or-below the
            # insert point -- A1:A2 becomes A2:A3, not A1:A3.
            vm.insert_rows(1, amount=1)
            self.assertEqual(vm.autofilter()["ref"], "A2:A3")
            out = os.path.join(d, "af_shift_out.xlsx")
            vm.save_workbook(out)
            vm_reload = elixcee.load_workbook(out)
            self.assertEqual(vm_reload.autofilter()["ref"], "A2:A3")

            # Second save-reload cycle must not drift.
            out2 = os.path.join(d, "af_shift_out2.xlsx")
            vm_reload.save_workbook(out2)
            vm_reload2 = elixcee.load_workbook(out2)
            self.assertEqual(vm_reload2.autofilter()["ref"], "A2:A3")

    def test_shared_mutation_safety_filtering_never_touches_cell_values(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["a", "b"])
        ws.append(["x", "y"])
        with tempfile.TemporaryDirectory() as d:
            src = os.path.join(d, "af_shared_src.xlsx")
            wb.save(src)
            vm = elixcee.load_workbook(src)
            vm.add_autofilter("A1:B2")
            vm.set_equality_filter(0, ["x"])
            before = vm.get_range("A1:B2")
            vm.set_equality_filter(0, ["nonexistent"])
            after = vm.get_range("A1:B2")
            self.assertEqual(before, after)


class TableAutoFilterColumnsAgreeWithOpenpyxl(unittest.TestCase):
    # 0.16.0-B2: set_table_equality_filter()/set_table_custom_filter()/
    # set_table_blank_filter()/set_table_top10_filter()/set_table_date_group_filter()/
    # clear_table_filter_column() -- the table-embedded mirror of
    # AutoFilterFilteringAgreesWithOpenpyxl's own standalone-autoFilter methods, reusing
    # the exact same FilterColumn/FilterCriteria model and row-hide evaluation engine.
    # fixture3's own real table (テーブル1, columns Name/Qty/Status, real data
    # gweg/33/good, wf/44/ok, fwf/55/bad) is real-fixture grounding for the common
    # equality case; the other filter types use openpyxl-synthetic fixtures, same
    # granted exception as the standalone case.

    def test_equality_filter_on_a_real_fixture_table_hides_non_matching_rows(self):
        vm = elixcee.load_workbook(FIXTURE3)
        t = vm.tables()[0]
        self.assertEqual(t["autofilter_columns"], [])
        self.assertEqual(vm.hidden_rows(), [])

        vm.set_table_equality_filter(t["name"], 1, ["44"])  # Qty == 44 -> only "wf" row
        self.assertEqual(
            vm.tables()[0]["autofilter_columns"],
            [{"col_offset": 1, "type": "values", "values": ["44"]}],
        )
        self.assertEqual(vm.hidden_rows(), [2, 4])  # rows for Qty=33 and Qty=55

        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "table_af_eq_out.xlsx")
            vm.save_workbook(out)
            xml = zipfile.ZipFile(out).read("xl/tables/table1.xml").decode()
            self.assertIn('<filterColumn colId="1">', xml)
            self.assertIn('<filter val="44"/>', xml)
            self.assertIsNotNone(openpyxl.load_workbook(out))

            # Second save cycle: criteria and GUIDs both stable.
            vm2 = elixcee.load_workbook(out)
            self.assertEqual(
                vm2.tables()[0]["autofilter_columns"],
                [{"col_offset": 1, "type": "values", "values": ["44"]}],
            )
            out2 = os.path.join(d, "table_af_eq_out2.xlsx")
            vm2.save_workbook(out2)
            xml2 = zipfile.ZipFile(out2).read("xl/tables/table1.xml").decode()
            self.assertIn('<filter val="44"/>', xml2)
            self.assertIn("xr:uid=", xml2)  # table's own GUID survives, not reserialized

            # Clear reverts to a bare, self-closing <autoFilter>, GUID intact.
            vm2.clear_table_filter_column(t["name"], 1)
            out3 = os.path.join(d, "table_af_eq_out3.xlsx")
            vm2.save_workbook(out3)
            xml3 = zipfile.ZipFile(out3).read("xl/tables/table1.xml").decode()
            self.assertNotIn("filterColumn", xml3)
            self.assertIn("<autoFilter", xml3)
            self.assertIn("xr:uid=", xml3)

    def test_custom_filter_and_top10_reuse_the_same_engine_as_standalone(self):
        # Built via elixcee's own create_table rather than an openpyxl-authored one:
        # openpyxl's Table() writes an ABSOLUTE relationship Target
        # ("/xl/tables/table1.xml") rather than this project's relative convention
        # ("../tables/table1.xml", confirmed against fixture3's real Excel-authored
        # file and matched by create_table's own A3 design) -- elixcee's table loader
        # doesn't resolve the absolute form (a real, pre-existing gap, unrelated to and
        # out of scope for 0.16.0-B2). Matches TablesAgreeWithOpenpyxl's own
        # from-scratch pattern; still verified via a real openpyxl reopen below.
        vm = elixcee.Vm()
        vm.set_cell(1, 1, "Name")
        vm.set_cell(1, 2, "Num")
        for i, (name, n) in enumerate([("A", 1), ("B", 5), ("C", 10), ("D", 15), ("E", 20)]):
            vm.set_cell(2 + i, 1, name)
            vm.set_cell(2 + i, 2, n)
        vm.create_table("A1:B6", name="T1")

        vm.set_table_custom_filter(
            "T1", 1, "greaterThanOrEqual", "10", and_=True, operator2="lessThanOrEqual", value2="15"
        )
        self.assertEqual(vm.hidden_rows(), [2, 3, 6])  # 1, 5, 20 fail; 10, 15 pass

        # Replacing column 1's own criteria (top10 instead of custom) means only the
        # top10 rule is now active for that column -- top 2 of [1,5,10,15,20] is
        # {15, 20}, so rows 2 ("A"/1), 3 ("B"/5), 4 ("C"/10) hide.
        vm.set_table_top10_filter("T1", 1, 2.0, top=True)
        self.assertEqual(vm.hidden_rows(), [2, 3, 4])

        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "table_af_custom_out.xlsx")
            vm.save_workbook(out)
            self.assertIsNotNone(openpyxl.load_workbook(out))
            xml = zipfile.ZipFile(out).read("xl/tables/table1.xml").decode()
            self.assertIn("<top10", xml)

    def test_blank_and_date_group_filters_round_trip(self):
        # Excel serials for 2024-01-15/2024-02-01/2024-01-20, confirmed against
        # openpyxl.utils.datetime.to_excel -- date_group_matches (src/vm/mod.rs)
        # decomposes a cell's raw numeric value via serial_to_ymd, it never accepts a
        # Python datetime object directly.
        vm = elixcee.Vm()
        vm.set_cell(1, 1, "X")
        vm.set_cell(1, 2, "D")
        vm.set_cell(2, 1, "a")
        vm.set_cell(2, 2, 45306.0)  # 2024-01-15
        vm.set_cell(3, 2, 45323.0)  # 2024-02-01, row 3's "X" cell left blank
        vm.set_cell(4, 1, "b")
        vm.set_cell(4, 2, 45311.0)  # 2024-01-20
        vm.create_table("A1:B4", name="T2")

        vm.set_table_blank_filter("T2", 0)
        self.assertEqual(vm.hidden_rows(), [2, 4])  # only the blank "X" row (3) stays

        vm.clear_table_filter_column("T2", 0)
        vm.set_table_date_group_filter("T2", 1, year=2024, month=1, grouping="month")
        self.assertEqual(vm.hidden_rows(), [3])  # only Feb (row 3) fails January

        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "table_af_date_out.xlsx")
            vm.save_workbook(out)
            self.assertIsNotNone(openpyxl.load_workbook(out))
            xml = zipfile.ZipFile(out).read("xl/tables/table1.xml").decode()
            self.assertIn("dateGroupItem", xml)
            self.assertIn('month="1"', xml)

    def test_unknown_table_name_and_out_of_range_offset_raise(self):
        vm = elixcee.Vm()
        vm.set_cell(1, 1, "A")
        vm.set_cell(2, 1, "x")
        vm.create_table("A1:A2", name="T3")
        with self.assertRaises(ValueError):
            vm.set_table_equality_filter("NoSuchTable", 0, ["x"])
        with self.assertRaises(ValueError):
            vm.set_table_equality_filter("T3", 5, ["x"])
        # Confirms the table itself DID load correctly -- otherwise both asserts above
        # would trivially pass for the wrong reason ("sheet has no tables" is also a
        # ValueError).
        vm.set_table_equality_filter("T3", 0, ["x"])
        self.assertEqual(vm.hidden_rows(), [])

    def test_a_standalone_filter_call_on_a_table_only_sheet_is_independent(self):
        # Confirms the two storage contexts (standalone autoFilter vs. a table's own
        # nested one) never get confused for each other.
        vm = elixcee.Vm()
        vm.set_cell(1, 1, "A")
        vm.set_cell(2, 1, "x")
        vm.create_table("A1:A2", name="T4")
        with self.assertRaises(ValueError):
            vm.set_equality_filter(0, ["x"])  # no standalone autoFilter on this sheet
        vm.set_table_equality_filter("T4", 0, ["x"])  # the table's own works fine
        self.assertEqual(vm.hidden_rows(), [])


class SortRangeAndMergeCellsRejectOversizedOrInvalidInput(unittest.TestCase):
    # Pins the PyO3-layer bound checks that have no Rust unit test of their
    # own (they live in #[cfg(feature = "python")] glue, not Vm-core logic) --
    # an oversized address here would otherwise write real geometry spanning
    # the whole sheet into the saved file, unlike get_range/iter_rows where
    # the only cost of an oversized address is a large allocation.
    def test_sort_range_rejects_an_address_beyond_the_sheet_bounds(self):
        vm = elixcee.Vm()
        with self.assertRaises(ValueError):
            vm.sort_range("A1:A1048577", key_col=1)

    def test_merge_cells_rejects_an_address_beyond_the_sheet_bounds(self):
        vm = elixcee.Vm()
        with self.assertRaises(ValueError):
            vm.merge_cells("A1:XFE1")

    def test_sort_range_rejects_a_key_col_outside_the_ranges_own_span(self):
        vm = elixcee.Vm()
        vm.set_range("A1:B2", [[1, 2], [3, 4]])
        with self.assertRaises(ValueError):
            vm.sort_range("A1:B2", key_col=3)


if __name__ == "__main__":
    unittest.main(verbosity=2)
