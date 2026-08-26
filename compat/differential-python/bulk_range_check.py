#!/usr/bin/env python3
"""Differential check: elixcee's bulk worksheet range/row API (R1) vs. openpyxl.

openpyxl is used PURELY as a test-only oracle here -- never a runtime dependency
of the shipped `elixcee` package (see pyproject.toml, which declares none).
Requires a `maturin develop --features python` build of elixcee and
`pip install openpyxl` in the active environment; see this directory's
README.md for the one-time setup.

Compares get_range()/iter_rows()/iter_cols() cell values against openpyxl's own
read of the same real fixture, checks append_row() on a sparse sheet against
openpyxl's own max_row after a round trip, and exercises the one true
end-to-end "no partial write on validation failure" case that only a real
PyO3 boundary (not a pure-Rust unit test) can prove.

One real, expected divergence is asserted rather than silently matched:
elixcee's max_row/max_column/calculate_dimension are a bounding box over cells
that actually hold a value; openpyxl's ws.dimensions mirrors the real XLSX
<dimension> element, which Excel widens to cover a merged range's full span
even when only the merge's anchor cell holds a value. fixture1's B1:C1 merge
makes this concrete: openpyxl reports "A1:C3", elixcee reports "A1:A3" (only
column A actually has values) -- see docs/openpyxl-gap-audit.md's
"Implementation notes for R1" for the underlying `Variant::Empty`-exclusion
convention this follows.

Run standalone: `python3 compat/differential-python/bulk_range_check.py`
"""

from __future__ import annotations

import os
import sys
import tempfile
import unittest

import elixcee
import openpyxl

FIXTURE = os.path.join(
    os.path.dirname(__file__),
    "..",
    "oracle-excel-com",
    "fixtures",
    "pristine",
    "fixture1_values_styles_merge_hidden.xlsm",
)


class GetRangeAndIterRowsAgreeWithOpenpyxl(unittest.TestCase):
    def setUp(self):
        self.vm = elixcee.load_workbook(FIXTURE)
        self.wb = openpyxl.load_workbook(FIXTURE, data_only=True)
        self.ws = self.wb.active

    def test_get_range_matches_openpyxl_over_the_populated_columns(self):
        # Only column A actually holds values in this fixture (B1:C1 is a
        # merge with no value of its own) -- compare column A only, since
        # elixcee's own get_range("A1:C3") and openpyxl's read of the same
        # range already agree cell-for-cell regardless (both report None for
        # B1/C1), this just narrows the assertion to what's meaningful.
        elixcee_values = [row[0] for row in self.vm.get_range("A1:A3")]
        openpyxl_values = [self.ws.cell(row=r, column=1).value for r in (1, 2, 3)]
        self.assertEqual(elixcee_values, openpyxl_values)

    def test_get_range_agrees_on_empty_merged_cells_too(self):
        elixcee_values = self.vm.get_range("A1:C3")
        openpyxl_values = [
            [self.ws.cell(row=r, column=c).value for c in (1, 2, 3)] for r in (1, 2, 3)
        ]
        self.assertEqual(elixcee_values, openpyxl_values)

    def test_iter_rows_matches_openpyxl_values_only_iteration(self):
        elixcee_rows = self.vm.iter_rows(max_col=3)
        openpyxl_rows = [
            [c.value for c in row] for row in self.ws.iter_rows(max_col=3)
        ]
        self.assertEqual(elixcee_rows, openpyxl_rows)

    def test_iter_cols_matches_openpyxl_values_only_iteration(self):
        elixcee_cols = self.vm.iter_cols(max_col=3)
        openpyxl_cols = [
            [c.value for c in col] for col in self.ws.iter_cols(max_col=3)
        ]
        self.assertEqual(elixcee_cols, openpyxl_cols)

    def test_dimension_divergence_from_a_merged_range_is_the_documented_one(self):
        # NOT an assertion of agreement -- pins the one real, expected
        # divergence documented in docs/openpyxl-gap-audit.md so a future
        # change to either side's convention is caught here, not silently.
        self.assertEqual(self.ws.dimensions, "A1:C3")
        self.assertEqual(self.vm.calculate_dimension(), "A1:A3")
        self.assertEqual((self.ws.max_row, self.ws.max_column), (3, 3))
        self.assertEqual((self.vm.max_row(), self.vm.max_column()), (3, 1))


class AppendRowMatchesOpenpyxlAfterARoundTrip(unittest.TestCase):
    def test_append_row_on_a_sparse_sheet_matches_openpyxls_own_max_row(self):
        # Loads the real fixture rather than a bare Vm() specifically so the
        # saved output carries real, passthrough-preserved styles.xml --
        # save_xlsx_impl's from-scratch stylesheet for a Vm with no loaded
        # file emits a bare empty <fill/> (no <patternFill> child), which
        # openpyxl's own reader rejects (TypeError: expected Fill) on
        # reopen. Pre-existing, unrelated to append_row itself (any
        # from-scratch Vm().save_workbook() hits it); out of scope for R1,
        # not fixed here, so the test routes around it instead of failing on
        # an orthogonal styles.xml gap.
        vm = elixcee.load_workbook(FIXTURE)
        written_row = vm.append_row(["x", "y"])
        self.assertEqual(written_row, 4)  # fixture1's used range ends at row 3

        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "out.xlsx")
            vm.save_workbook(path)
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            self.assertEqual(ws.max_row, 4)
            self.assertEqual([ws.cell(row=4, column=c).value for c in (1, 2)], ["x", "y"])


class SetRangeHasNoPartialWriteOnValidationFailure(unittest.TestCase):
    def test_a_bad_typed_element_not_at_index_zero_leaves_the_range_untouched(self):
        vm = elixcee.Vm()
        vm.set_range("A1:B1", [[1, 2]])
        before = vm.get_range("A1:B1")

        with self.assertRaises(TypeError):
            # The bad element (a dict, unsupported by py_to_variant) is at
            # index 1, not 0 -- if set_range wrote as it went instead of
            # validating the whole grid first, A1 would already be
            # overwritten by the time this raises.
            vm.set_range("A1:B1", [[99, {"bad": "type"}]])

        after = vm.get_range("A1:B1")
        self.assertEqual(before, after)
        self.assertEqual(after, [[1, 2]])


class MaxRowMaxColumnCalculateDimensionOnAnEmptySheet(unittest.TestCase):
    def test_all_three_are_none_on_a_genuinely_empty_sheet(self):
        vm = elixcee.Vm()
        self.assertIsNone(vm.max_row())
        self.assertIsNone(vm.max_column())
        self.assertIsNone(vm.calculate_dimension())

    def test_openpyxl_reports_a1_a1_for_the_same_empty_sheet(self):
        # Registered divergence, not a bug in either library: openpyxl's
        # brand-new-worksheet convention is "A1" for max_row/max_column == 1;
        # elixcee's is None, chosen so a caller can distinguish "empty" from
        # "one real value at A1" without a separate emptiness check.
        wb = openpyxl.Workbook()
        ws = wb.active
        self.assertEqual((ws.max_row, ws.max_column), (1, 1))


if __name__ == "__main__":
    unittest.main(verbosity=2)
